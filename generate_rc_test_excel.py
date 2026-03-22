"""
Generate Excel workbook to verify Regional Controller (RcTrends) dashboard logic.

Sheets produced:
  1. Submissions          -- input: ~1,400 fake daily submission rows (5 locations x 12 months)
  2. Monthly_All          -- AVERAGEIFS per section per month, all locations combined
  3. Monthly_APPLETON      -- same, filtered to APPLETON only
  4. Monthly_CENTRAL_IL    -- CENTRAL IL
  5. Monthly_DES_MOINES    -- DES MOINES
  6. Monthly_MADISON       -- MADISON
  7. Monthly_OMAHA         -- OMAHA
  8. Weekly_All           -- AVERAGEIFS per ISO week
  9. Quarterly_All        -- AVERAGEIFS per quarter
 10. KPI_Summary          -- Latest / Prev / %Change / PeriodAvg / Peak for every section
 11. Logic_Map            -- documents which dashboard value = which formula
"""

import random
import math
from datetime import date, timedelta

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

random.seed(42)

# ── Config (mirrors RcTrends.tsx SECTIONS) ────────────────────────────────────
SECTIONS = [
    ('A', 'Currency Bills', 6100),
    ('B', 'Loose Coins',     470),
    ('C', 'Machine Coin',    760),
    ('D', 'Rolled Coins',    920),
    ('E', 'Checks',          460),
    ('F', 'Other',           170),
    ('G', 'Transfers',       370),
    ('H', 'Petty Cash',      165),
    ('I', 'Safe Loan',       160),
]

# mirrors RcTrends.tsx LOC_MULT (approximate; uses actual IDs from seeded DB)
LOCATIONS = [
    ('loc-appleton',   'APPLETON',    1.00),
    ('loc-central-il', 'CENTRAL IL',  0.94),
    ('loc-des-moines', 'DES MOINES',  0.97),
    ('loc-madison',    'MADISON',     0.88),
    ('loc-omaha',      'OMAHA',       0.82),
]

TODAY    = date(2026, 3, 6)
START    = date(2025, 3, 7)   # ~12 months of history

# ── Styling helpers ───────────────────────────────────────────────────────────
HDR_GREEN  = PatternFill('solid', fgColor='1A5C38')
HDR_FONT   = Font(bold=True, color='FFFFFF', size=10)
NOTE_FONT  = Font(italic=True, color='595959', size=9)
TITLE_FONT = Font(bold=True, color='1A5C38', size=12)
GRAY_FILL  = PatternFill('solid', fgColor='F2F2F2')
NUM_FMT    = '#,##0.00'
PCT_FMT    = '0.00"%"'

def hdr(cell, text, fill=None):
    cell.value = cell.value if text is None else text
    cell.fill  = fill or HDR_GREEN
    cell.font  = HDR_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center')

def fmt_num(ws, col_letter, row_start, row_end):
    for r in range(row_start, row_end + 1):
        ws[f'{col_letter}{r}'].number_format = NUM_FMT

def set_col_width(ws, mapping):
    """mapping: {col_letter: width}"""
    for col, w in mapping.items():
        ws.column_dimensions[col].width = w


# ─────────────────────────────────────────────────────────────────────────────
# STEP 1 — Generate fake submission rows
# ─────────────────────────────────────────────────────────────────────────────
rows = []   # list of lists matching Submissions header
sub_id = 1
cur = START
while cur <= TODAY:
    if cur.weekday() < 5:                         # weekdays only
        for loc_id, loc_name, mult in LOCATIONS:
            if random.random() > 0.12:            # ~12% absence/closed days
                m = cur.month
                # Seasonal factor (mirrors dashboard mock logic)
                seasonal = 0.96 if m in (6, 7, 8) else 1.05 if m == 12 else 1.0
                # Shared location-level noise (mild)
                loc_noise = random.gauss(0, 0.05)
                sec_vals = []
                total = 0.0
                for sec_code, _, base in SECTIONS:
                    # Each section gets slight independent noise on top
                    val = round(base * mult * seasonal * (1 + loc_noise + random.gauss(0, 0.03)), 2)
                    val = max(0.0, val)
                    sec_vals.append(val)
                    total += val
                rows.append([
                    f'SUB-{sub_id:04d}',          # A  Submission ID
                    loc_id,                         # B  Location ID
                    loc_name,                       # C  Location Name
                    cur.isoformat(),                # D  Date (YYYY-MM-DD)
                    cur.strftime('%Y-%m'),          # E  Month-Year
                    f"{cur.year}-W{cur.isocalendar()[1]:02d}",  # F  ISO Week
                    f"{cur.year}-Q{(cur.month-1)//3+1}",       # G  Quarter
                    'approved',                     # H  Status
                    *sec_vals,                      # I..Q  Sec A-I
                    round(total, 2),                # R  Total Cash
                ])
                sub_id += 1
    cur += timedelta(days=1)

print(f"Generated {len(rows)} submission rows")


# ─────────────────────────────────────────────────────────────────────────────
# STEP 2 — Build period lists
# ─────────────────────────────────────────────────────────────────────────────
# Monthly periods
months = []
d = date(START.year, START.month, 1)
while d <= TODAY:
    months.append(d.strftime('%Y-%m'))
    d = date(d.year + (1 if d.month == 12 else 0), 1 if d.month == 12 else d.month + 1, 1)

# ISO weeks
weeks_set = set()
d = START
while d <= TODAY:
    weeks_set.add(f"{d.year}-W{d.isocalendar()[1]:02d}")
    d += timedelta(days=7)
weeks = sorted(weeks_set)

# Quarters
quarters, seen_q = [], set()
d = date(START.year, START.month, 1)
while d <= TODAY:
    q = f"{d.year}-Q{(d.month-1)//3+1}"
    if q not in seen_q:
        quarters.append(q)
        seen_q.add(q)
    next_q_month = ((d.month - 1) // 3 + 1) * 3 + 1
    d = date(d.year + (1 if next_q_month > 12 else 0), (next_q_month - 1) % 12 + 1, 1)

print(f"Months: {months[0]} -> {months[-1]}  ({len(months)} periods)")
print(f"Weeks : {weeks[0]}  -> {weeks[-1]}  ({len(weeks)} periods)")
print(f"Qtrs  : {quarters[0]}  -> {quarters[-1]}  ({len(quarters)} periods)")


# ─────────────────────────────────────────────────────────────────────────────
# STEP 3 — Build workbook
# ─────────────────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ── Sheet 1: Submissions ──────────────────────────────────────────────────────
ws_sub = wb.active
ws_sub.title = 'Submissions'

SUB_HEADERS = [
    'Submission ID', 'Location ID', 'Location Name', 'Date', 'Month-Year',
    'ISO Week', 'Quarter', 'Status',
    'Sec A — Currency Bills', 'Sec B — Loose Coins', 'Sec C — Machine Coin',
    'Sec D — Rolled Coins',   'Sec E — Checks',      'Sec F — Other',
    'Sec G — Transfers',      'Sec H — Petty Cash',  'Sec I — Safe Loan',
    'Total Cash',
]
ws_sub.append(SUB_HEADERS)
for cell in ws_sub[1]:
    hdr(cell, None)

for row in rows:
    ws_sub.append(row)

# Number format for currency cols I..R
last_sub_row = len(rows) + 1
for col_letter in [get_column_letter(c) for c in range(9, 19)]:   # I to R
    fmt_num(ws_sub, col_letter, 2, last_sub_row)

# Column widths
sub_widths = {
    'A': 12, 'B': 16, 'C': 14, 'D': 12, 'E': 12,
    'F': 12,  'G': 10, 'H': 10,
    **{get_column_letter(c): 20 for c in range(9, 18)},
    'R': 14,
}
set_col_width(ws_sub, sub_widths)
ws_sub.freeze_panes = 'A2'
ws_sub.row_dimensions[1].height = 30

# Excel table
tbl = Table(
    displayName='Submissions',
    ref=f'A1:{get_column_letter(len(SUB_HEADERS))}{last_sub_row}'
)
tbl.tableStyleInfo = TableStyleInfo(name='TableStyleMedium7', showRowStripes=True)
ws_sub.add_table(tbl)


# ─────────────────────────────────────────────────────────────────────────────
# Helper: write a trend sheet with AVERAGEIFS formulas
# ─────────────────────────────────────────────────────────────────────────────
# Columns in Submissions sheet:
#   I=SecA  J=SecB  K=SecC  L=SecD  M=SecE  N=SecF  O=SecG  P=SecH  Q=SecI  R=Total
SEC_COLS    = ['I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q']  # 9 sections
TOTAL_COL   = 'R'

def write_trend_sheet(ws, periods, period_col, sheet_title, loc_filter=None):
    """
    periods     : list of period strings e.g. ['2026-01', '2026-02', ...]
    period_col  : column letter in Submissions that holds the period key (E=month, F=week, G=quarter)
    sheet_title : shown in row 1
    loc_filter  : location_id string to add a second AVERAGEIFS criterion, or None for all
    """
    # Row 1 — title
    ws['A1'] = sheet_title
    ws['A1'].font  = TITLE_FONT
    ws['A1'].alignment = Alignment(vertical='center')
    ws.row_dimensions[1].height = 22
    ws.merge_cells(f'A1:{get_column_letter(len(SECTIONS) + 3)}1')

    # Row 2 — note explaining the formula
    if loc_filter:
        note = (f'Formula: =AVERAGEIFS(Submissions!<sec_col>:<sec_col>, '
                f'Submissions!{period_col}:{period_col}, "<period>", '
                f'Submissions!B:B, "{loc_filter}")')
    else:
        note = (f'Formula: =AVERAGEIFS(Submissions!<sec_col>:<sec_col>, '
                f'Submissions!{period_col}:{period_col}, "<period>")')
    ws['A2'] = note
    ws['A2'].font = NOTE_FONT
    ws.merge_cells(f'A2:{get_column_letter(len(SECTIONS) + 3)}2')

    # Row 3 — column headers
    hdr_row = ['Period'] + [f'Avg {s[0]} — {s[1]}' for s in SECTIONS] + ['Count', 'Avg Total Cash']
    ws.append(hdr_row)
    for cell in ws[3]:
        hdr(cell, None)
    ws.row_dimensions[3].height = 28

    # Rows 4+ — one row per period
    for p_idx, period in enumerate(periods):
        row_cells = [period]

        for ci, (sec_code, _, _) in enumerate(SECTIONS):
            sc = SEC_COLS[ci]
            if loc_filter:
                f = (f'=IFERROR(AVERAGEIFS(Submissions!${sc}:${sc},'
                     f'Submissions!${period_col}:${period_col},"{period}",'
                     f'Submissions!$B:$B,"{loc_filter}"),"")')
            else:
                f = (f'=IFERROR(AVERAGEIFS(Submissions!${sc}:${sc},'
                     f'Submissions!${period_col}:${period_col},"{period}"),"")')
            row_cells.append(f)

        # Count of submissions in this period
        if loc_filter:
            count_f = (f'=COUNTIFS(Submissions!${period_col}:${period_col},"{period}",'
                       f'Submissions!$B:$B,"{loc_filter}")')
        else:
            count_f = f'=COUNTIFS(Submissions!${period_col}:${period_col},"{period}")'
        row_cells.append(count_f)

        # Avg total cash
        if loc_filter:
            total_f = (f'=IFERROR(AVERAGEIFS(Submissions!${TOTAL_COL}:${TOTAL_COL},'
                       f'Submissions!${period_col}:${period_col},"{period}",'
                       f'Submissions!$B:$B,"{loc_filter}"),"")')
        else:
            total_f = (f'=IFERROR(AVERAGEIFS(Submissions!${TOTAL_COL}:${TOTAL_COL},'
                       f'Submissions!${period_col}:${period_col},"{period}"),"")')
        row_cells.append(total_f)

        ws.append(row_cells)

    # Number format for all value columns
    data_end = 3 + len(periods)
    for ci in range(1, len(SECTIONS) + 3):   # skip Period col (col 0)
        col_ltr = get_column_letter(ci + 1)
        for r in range(4, data_end + 1):
            ws[f'{col_ltr}{r}'].number_format = NUM_FMT

    # Alternating row shading
    for r in range(4, data_end + 1, 2):
        for c in range(1, len(SECTIONS) + 4):
            ws.cell(r, c).fill = GRAY_FILL

    # Column widths
    ws.column_dimensions['A'].width = 13
    for ci in range(2, len(SECTIONS) + 4):
        ws.column_dimensions[get_column_letter(ci)].width = 19

    ws.freeze_panes = 'B4'
    ws.row_dimensions[3].height = 30

    return 3 + len(periods)   # last data row index (1-based)


# ── Sheet 2: Monthly_All ──────────────────────────────────────────────────────
ws_mo_all = wb.create_sheet('Monthly_All')
mo_last_row = write_trend_sheet(
    ws_mo_all, months, 'E',
    'Monthly Averages — ALL Locations  |  mirrors GET /reports/section-trends?granularity=monthly'
)

# ── Sheets 3-7: Monthly per location ─────────────────────────────────────────
for loc_id, loc_name, _ in LOCATIONS:
    safe_name = f'Monthly_{loc_name.replace(" ", "_")[:11]}'
    ws_loc = wb.create_sheet(safe_name)
    write_trend_sheet(
        ws_loc, months, 'E',
        f'Monthly Averages — {loc_name}  |  mirrors ?location_id={loc_id}',
        loc_filter=loc_id
    )

# ── Sheet 8: Weekly_All ───────────────────────────────────────────────────────
ws_wk_all = wb.create_sheet('Weekly_All')
write_trend_sheet(
    ws_wk_all, weeks, 'F',
    'Weekly Averages — ALL Locations  |  mirrors GET /reports/section-trends?granularity=weekly'
)

# ── Sheet 9: Quarterly_All ────────────────────────────────────────────────────
ws_qt_all = wb.create_sheet('Quarterly_All')
write_trend_sheet(
    ws_qt_all, quarters, 'G',
    'Quarterly Averages — ALL Locations  |  mirrors GET /reports/section-trends?granularity=quarterly'
)


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 10: KPI_Summary  (mirrors the 4 KPI cards in the dashboard)
#
# In Monthly_All:
#   Row 1  = title          Row 2 = note
#   Row 3  = header         Rows 4..(3+len(months)) = period data
#
# Columns in Monthly_All: A=Period, B=SecA, C=SecB, ..., J=SecI, K=Count, L=TotalCash
# ─────────────────────────────────────────────────────────────────────────────
ws_kpi = wb.create_sheet('KPI_Summary')

last_period_row = 3 + len(months)   # last data row in Monthly_All
prev_period_row = last_period_row - 1

# Title
ws_kpi['A1'] = 'DASHBOARD KPI SUMMARY — Regional Controller (RcTrends.tsx)'
ws_kpi['A1'].font = Font(bold=True, size=13, color='1A5C38')
ws_kpi.merge_cells('A1:H1')
ws_kpi.row_dimensions[1].height = 26

# Explanation block
notes = [
    '',
    'How this sheet works:',
    '  Each row = one section (A-I).  All KPI values are Excel formulas that pull from Monthly_All.',
    '  This mirrors exactly what the RcTrends dashboard shows for the DEFAULT view (Monthly, 6 periods, All locations).',
    '',
    '  Latest     = average for the most recent month   -->  RcTrends.tsx: const latest = values.at(-1)',
    '  Previous   = average for the month before that   -->  RcTrends.tsx: const prev   = values.at(-2)',
    '  % Change   = (Latest - Previous) / Previous * 100  -->  RcTrends.tsx: pctChange  = (latest-prev)/prev*100',
    '  Period Avg = AVERAGE of all monthly averages     -->  RcTrends.tsx: Math.round(sum/length)',
    '  Peak       = MAX  of all monthly averages        -->  RcTrends.tsx: Math.max(...values)',
    '',
    '  Backend (reports.py): avg_total = round(sum(bucket[period]) / len(bucket[period]), 2)',
    '  This is AVERAGEIFS in Excel.  The Monthly_All sheet implements this for every period.',
    '',
]
for i, n in enumerate(notes, start=2):
    ws_kpi.cell(i, 1, n).font = NOTE_FONT
    if n and not n.startswith(' '):
        ws_kpi.cell(i, 1).font = Font(bold=True, size=10, color='1A5C38')

header_row = 2 + len(notes)

# Column headers for the KPI table
kpi_cols = ['Section', 'Name', 'Latest (last month)', 'Previous month',
            '% Change', 'Period Average (all months)', 'Peak', 'Trend']
ws_kpi.append([''] * 1)   # blank spacer — move to exact row
# Rewrite at correct row
for ci, col_name in enumerate(kpi_cols, start=1):
    cell = ws_kpi.cell(header_row, ci, col_name)
    hdr(cell, None)
ws_kpi.row_dimensions[header_row].height = 28

# Data rows — one per section
# Monthly_All column layout: A=Period, B=SecA, ..., J=SecI
sec_col_letters_mo = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']

for si, (sec_code, sec_name, _) in enumerate(SECTIONS):
    sc = sec_col_letters_mo[si]
    data_range = f"Monthly_All!${sc}$4:Monthly_All!${sc}${last_period_row}"
    latest_ref = f"Monthly_All!${sc}${last_period_row}"
    prev_ref   = f"Monthly_All!${sc}${prev_period_row}"

    latest_f  = f'={latest_ref}'
    prev_f    = f'={prev_ref}'
    pct_f     = f'=IFERROR(({latest_ref}-{prev_ref})/{prev_ref}*100,0)'
    avg_f     = f'=IFERROR(AVERAGE({data_range}),0)'
    peak_f    = f'=IFERROR(MAX({data_range}),0)'
    trend_f   = f'=IF({latest_ref}="","",IF(({latest_ref}-{prev_ref})>0,"UP","DOWN"))'

    r = header_row + 1 + si
    ws_kpi.cell(r, 1, sec_code)
    ws_kpi.cell(r, 2, sec_name)
    ws_kpi.cell(r, 3, latest_f).number_format  = NUM_FMT
    ws_kpi.cell(r, 4, prev_f).number_format    = NUM_FMT
    ws_kpi.cell(r, 5, pct_f).number_format     = '0.00"%"'
    ws_kpi.cell(r, 6, avg_f).number_format     = NUM_FMT
    ws_kpi.cell(r, 7, peak_f).number_format    = NUM_FMT
    ws_kpi.cell(r, 8, trend_f)

# Column widths
kpi_widths = {'A': 10, 'B': 18, 'C': 22, 'D': 18, 'E': 14, 'F': 28, 'G': 18, 'H': 10}
set_col_width(ws_kpi, kpi_widths)


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 11: Logic_Map — formula traceability
# ─────────────────────────────────────────────────────────────────────────────
ws_lm = wb.create_sheet('Logic_Map')
ws_lm['A1'] = 'FORMULA TRACEABILITY — Dashboard Element → Backend Code → Frontend Code → Excel Formula'
ws_lm['A1'].font = TITLE_FONT
ws_lm.merge_cells('A1:F1')
ws_lm.row_dimensions[1].height = 24

lm_headers = ['Dashboard Element', 'API Endpoint', 'Backend (reports.py)',
              'Frontend (RcTrends.tsx)', 'Excel Sheet', 'Excel Formula']
ws_lm.append(lm_headers)
for cell in ws_lm[2]:
    hdr(cell, None)
ws_lm.row_dimensions[2].height = 26

lm_rows = [
    [
        'Chart data points (Y axis)',
        'GET /reports/section-trends ?section=A &granularity=monthly &periods=6',
        'bucket[period].append(section["total"])\navg = round(sum/len, 2)',
        'chartData = apiTrends.data.map(p => ({period:p.period, [sectionKey]:p.avg_total}))',
        'Monthly_All  cols B-J',
        '=AVERAGEIFS(Submissions!$I:$I, Submissions!$E:$E, "2026-02")',
    ],
    [
        '"Latest" KPI card',
        'summary.latest_value',
        'values[-1]  (last period in ordered bucket)',
        'const latest = values.at(-1)',
        'KPI_Summary  col C',
        '=Monthly_All!$B$<last_row>',
    ],
    [
        '"Average" KPI card',
        'summary.period_avg',
        'mean of all period avg_totals',
        'Math.round(values.reduce((a,b)=>a+b,0) / values.length)',
        'KPI_Summary  col F',
        '=AVERAGE(Monthly_All!$B$4:$B$<last_row>)',
    ],
    [
        '"Peak" KPI card',
        'summary.peak',
        'max(values)',
        'Math.max(...values)',
        'KPI_Summary  col G',
        '=MAX(Monthly_All!$B$4:$B$<last_row>)',
    ],
    [
        '% Change label (up/down arrow)',
        'summary.change_pct',
        'round((latest-prev)/prev*100, 1)',
        'const pctChange = prev ? ((latest-prev)/prev)*100 : 0',
        'KPI_Summary  col E',
        '=IFERROR((latest_cell - prev_cell)/prev_cell*100, 0)',
    ],
    [
        'Location filter (pill buttons)',
        '?location_id=<id>  or omit for All',
        'q.filter(Submission.location_id == location_id)',
        'locationId === "all" ? undefined : locationId',
        'Monthly_<LOC> sheets',
        '=AVERAGEIFS(..., Submissions!$B:$B, "loc-appleton")',
    ],
    [
        'Period / X-axis label',
        'data[].period',
        '_period_key(): YYYY-MM | YYYY-Wnn | YYYY-Qn',
        'p.period',
        'Monthly_All col A / Weekly_All col A / Quarterly_All col A',
        'Static text in col A of each trend sheet',
    ],
    [
        'Granularity toggle (weekly/monthly/quarterly)',
        '?granularity=monthly',
        'granularity param selects _period_key() logic',
        'const [granularity, setGranularity] = useState("monthly")',
        'Weekly_All  /  Monthly_All  /  Quarterly_All',
        'Separate sheets, one per granularity',
    ],
    [
        'Count per period (tooltip)',
        'Not in API — frontend derives from chartData.length',
        'N/A',
        'across {chartData.length} {periodUnit}',
        'Monthly_All  col K',
        '=COUNTIFS(Submissions!$E:$E, "2026-02")',
    ],
    [
        'Mock/fallback data (no backend)',
        'N/A — frontend-only',
        'N/A',
        'monthlyPts() / weeklyPts() / quarterlyPts() using sin-based noise',
        'N/A — not replicated (requires JS Math.sin)',
        'Excel uses real AVERAGEIFS; mock data path is not testable here',
    ],
]

for row in lm_rows:
    ws_lm.append(row)
    # wrap text in all cells of this row
    r_num = ws_lm.max_row
    for c in range(1, 7):
        ws_lm.cell(r_num, c).alignment = Alignment(wrap_text=True, vertical='top')

# Alternating shading
for r in range(3, ws_lm.max_row + 1, 2):
    for c in range(1, 7):
        ws_lm.cell(r, c).fill = GRAY_FILL

lm_widths = {'A': 32, 'B': 40, 'C': 42, 'D': 50, 'E': 30, 'F': 55}
set_col_width(ws_lm, lm_widths)
ws_lm.freeze_panes = 'A3'
for r in range(3, ws_lm.max_row + 1):
    ws_lm.row_dimensions[r].height = 55


# ─────────────────────────────────────────────────────────────────────────────
# Save
# ─────────────────────────────────────────────────────────────────────────────
OUT = r'E:\Master - slave\Damco material\Compass\RC_Dashboard_Verification.xlsx'
wb.save(OUT)

print()
print(f'Saved -> {OUT}')
print()
print('Sheets in workbook:')
for s in wb.sheetnames:
    print(f'  {s}')
print()
print('How to use:')
print('  1. Open RC_Dashboard_Verification.xlsx')
print('  2. Check KPI_Summary sheet -> compare each row with the dashboard KPI cards')
print('     (select a section on dashboard, note Latest/Average/Peak/% Change)')
print('  3. Check Monthly_All -> compare chart data points period by period')
print('  4. Change any Sec value in Submissions tab -> KPIs and chart data update automatically')
print('  5. Logic_Map sheet explains every formula and its source in code')
